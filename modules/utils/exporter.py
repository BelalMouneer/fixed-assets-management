from openpyxl import Workbook
from modules.models.models import DailyEntries, Account
import datetime

def export_all_accounts_to_xlsx(session):
    wb = Workbook()
    
    # Remove the default sheet if it exists
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Get all column names from Accounts dynamically
    headers = list(Account.__table__.columns.keys())
    ws = wb.create_sheet(title="Accounts")
    ws.append(headers)

    # Query all accounts from the database
    for account in session.query(Account).all():
        row = []
        for col in headers:
            value = getattr(account, col)
            row.append(value)
        ws.append(row)

    return wb     

def export_all_entries_to_xlsx(session):
    wb = Workbook()
    
    # Remove the default sheet if it exists
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Get all column names from DailyEntries dynamically
    headers = list(DailyEntries.__table__.columns.keys())
    ws = wb.create_sheet(title="Daily Entries")
    ws.append(headers)
    
    # Query all daily entries from the database
    for entry in session.query(DailyEntries).all():
        row = []
        for col in headers:
            value = getattr(entry, col)
            if isinstance(value, (datetime.date, datetime.datetime)):
                # Use a safe format and remove any leading zero for month/day
                formatted = value.strftime("%m/%d/%Y/%H:%M:%S")
                # Remove leading zero from month and day
                formatted = formatted.lstrip("0").replace("/0", "/")
                value = formatted
            row.append(value)
        ws.append(row)
    
    return wb
